import openpyxl
from openpyxl import Workbook
import CopySheet
from pathlib import Path


def FormatarQlik(label, caminho):
    workbookQlik = openpyxl.load_workbook(caminho)
    sheet = workbookQlik.worksheets[0]
    colunasParaInserir = 3
    deletados = 0
    if sheet.max_column == 17:
        # apagar cod 6000+
        max_rows = sheet.max_row
        for row in range(sheet.max_row, 1, -1):
            label.config(
                text=f"Agagando serviços cod. 6000+ | Total de linhas: {sheet.max_row}\
linha processada: {row} (iniciou com {max_rows} linhas)"
            )
            if (
                sheet[f"L{row}"].value is not None
                and int(sheet[f"L{row}"].value) > 6000
            ):
                sheet.delete_rows(row)
                deletados += 1
        # insere colunas
        sheet.insert_cols(2, colunasParaInserir)
        # copia valores da coluna A para B, C e D
        for row in range(2, sheet.max_row + 1):
            label.config(
                text=f"Separando Data e hora dos protocolos: linha processada: {row} de {sheet.max_row}"
            )
            sheet[f"B{row}"] = sheet[f"A{row}"].value[:8]
            sheet[f"C{row}"].number_format = 'General'
            sheet[f"C{row}"] = int(sheet[f"A{row}"].value[8:12])
            sheet[f"D{row}"] = sheet[f"A{row}"].value[12:17]
    elif sheet.max_column == 20:
        # apagar cod 6000+
        max_rows = sheet.max_row
        for row in range(sheet.max_row, 1, -1):
            label.config(
                text=f"Agagando serviços cod. 6000+ | Total de linhas: {sheet.max_row}\
linha processada: {row} (iniciou com {max_rows} linhas)"
            )
            if (
                sheet[f"L{row}"].value is not None
                and int(sheet[f"L{row}"].value) > 6000
            ):
                sheet.delete_rows(row)
                deletados += 1
        # copia valores da coluna A para B, C e D
        for row in range(2, sheet.max_row + 1):
            label.config(
                text=f"Separando Data e hora dos protocolos: linha processada: {row} de {sheet.max_row}"
            )
            sheet[f"B{row}"] = sheet[f"A{row}"].value[:8]
            sheet[f"C{row}"].number_format = 'General'
            sheet[f"C{row}"] = int(sheet[f"A{row}"].value[8:12])
            sheet[f"D{row}"] = sheet[f"A{row}"].value[12:17]
    workbookQlik.save(caminho)
    workbookQlik.close()
    sheet["B1"] = "Data"
    sheet["C1"] = "Hora"
    sheet["D1"] = "Chave"
    workbookQlik.save(caminho)
    label.config(text=f"Qlik Concluido, {deletados} linhas deletadas")
    workbookQlik.close()


def FormatarAtrasados(label, caminho):
    workbookAtrasados = openpyxl.load_workbook(caminho)
    sheet = workbookAtrasados.worksheets[0]
    deletados = 0
    for row in range(sheet.max_row, 2, -1):
        label.config(
            text=f"Removendo expurgados e sem resposta: linha processada: {row} de {sheet.max_row}"
        )
        if (
            sheet[f"V{row}"].value is not None
            and sheet[f"V{row}"].value.lower() == "sim"
        ):
            sheet.delete_rows(row)
    workbookAtrasados.save(caminho)
    label.config(text=f"Serviços em atraso Concluido, {deletados} linhas deletadas")
    workbookAtrasados.close()


def FormatarBDOs(label, caminho):
    workbookBDOs = openpyxl.load_workbook(caminho)
    sheet = workbookBDOs.worksheets[0]
    deletados = 0
    for row in range(sheet.max_row, 3, -1):
        # esparso
        if (
            sheet[f"C{row}"].value is not None
            and sheet[f"C{row}"].value.lower() == "esparso"
        ):
            sheet.delete_rows(row)
            deletados += 1

        # expurgados
        if (
            sheet[f"T{row}"].value is not None
            and sheet[f"T{row}"].value.lower() == "sim"
        ):
            sheet.delete_rows(row)
            deletados += 1

    # BDO automatico
    for row in range(sheet.max_row, 3, -1):
        if sheet[f"O{row}"].value is not None and (
            "BDO AUTOMÁTICO - Protocolo 2461: " in sheet[f"O{row}"].value
        ):
            splited_text = str(sheet[f"O{row}"].value).split()
            sheet[f"O{row}"] = " ".join(splited_text[9:])
    workbookBDOs.save(caminho)
    label.config(text=f"BDOs Concluido, {deletados} linhas deletadas")
    workbookBDOs.close()


def copiarPlanilha(
    caminhoSalvar,
    targetWorkbook,
    targetWorksheet,
    sourceWorkbook,
    sourceWorksheet,
):
    my_file = Path(f"{caminhoSalvar}/{targetWorkbook}")
    if not (my_file.is_file()):
        workbook = Workbook()
        workbook.save(f"{caminhoSalvar}/{targetWorkbook}")
    else:
        workbook = openpyxl.load_workbook(f"{caminhoSalvar}/{targetWorkbook}")
    CopySheet.copysheet(
        sourceWorkbook=sourceWorkbook,
        targetWorkbook=my_file,
        worksheetSourceTitle=sourceWorksheet,
        worksheetTargetTitle=targetWorksheet,
    )
