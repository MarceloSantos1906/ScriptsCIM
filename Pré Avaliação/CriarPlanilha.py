from openpyxl.styles import *
import openpyxl
import locale
import time


def formating_row(sheet, row: int, size: float):
    sheet.row_dimensions[row].height = size


def formating_column(sheet, column: str, size: float):
    sheet.column_dimensions[column].width = size


def create_sheet(caminho):
    workbook = openpyxl.Workbook()
    workbook.save(f"{caminho}")
    locale.setlocale(locale.LC_TIME, "pt_BR")
    today = time.strftime("%B").upper()

    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook.create_sheet("Relatorio")

    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    ###     setting default row size
    for row in range(1000):
        formating_row(sheet, row, 20)

    ###     setting row sizes
    rows = [2, 3, 9, 14, 19, 23, 27]
    row_sizes = [7, 39.5, 7, 7, 7, 7, 7]
    for i in range(len(rows)):
        formating_row(sheet, rows[i], row_sizes[i])

    ###     setting column sizes
    columns = ["B", "C", "D", "E", "F", "G", "H"]
    column_sizes = [1, 27.14, 16.86, 15.14, 10.86, 15, 1]
    for i in range(len(columns)):
        formating_column(sheet, columns[i], column_sizes[i])

    ###     setting cell values
    cell_values = [
        [
            f"PRÉVIA DOS ITENS 1 E 2 - AVALIAÇÃO DO MÊS DE {today}",
            "INDICE DE PRAZO DE EXECUÇÃO - ITEM 1 A",
            "ITEM",
            "SGM - MANUTENÇÃO",
            "PAVIMENTOS",
            "SERVS DAS 22 ÀS 07",
            "",
            "INDICADOR DE RETRABALHO - ITEM 1 B",
            "ITEM",
            "SGM - MANUTENÇÃO",
            "PAVIMENTOS",
            "",
            "INDICADOR DE RETRABALHO - ITEM 1 C",
            "ITEM",
            "SGM - MANUTENÇÃO",
            "PAVIMENTOS",
            "",
            "INDICE DE PRAZO DE EXECUÇÃO - OBRAS - ITEM 2 A",
            "ITEM",
            "SGM - MANUTENÇÃO",
            "",
            "INDICE DE PRAZO DE EXECUÇÃO - OBRAS - ITEM 2 A",
            "ITEM",
            "SGM - MANUTENÇÃO",
            "",
        ],
        [
            "",
            "",
            "SERVS EXEC",
            0,
            0,
            0,
            "",
            "",
            "SERV FISCALIZADOS",
            0,
            0,
            "",
            "",
            "SERV FISCALIZADOS",
            0,
            0,
            "",
            "",
            "SERVS EXEC",
            0,
            "",
            "",
            "SERV FISCALIZADOS",
            0,
            "",
        ],
        [
            "",
            "",
            "EM ATRASO",
            0,
            0,
            0,
            "",
            "",
            "NÃO CONFORMES",
            0,
            0,
            "",
            "",
            "NÃO CONFORMES",
            0,
            0,
            "",
            "",
            "EM ATRASO",
            0,
            "",
            "",
            "NÃO CONFORMES",
            0,
            "",
        ],
        [
            "",
            "",
            "%",
            0,
            0,
            0,
            "",
            "",
            "%",
            0,
            0,
            "",
            "",
            "%",
            0,
            0,
            "",
            "",
            "%",
            0,
            "",
            "",
            "%",
            0,
            "",
        ],
        [
            "",
            "",
            "RESULTADO",
            0,
            0,
            0,
            "",
            "",
            "RESULTADO",
            0,
            0,
            "",
            "",
            "RESULTADO",
            0,
            0,
            "",
            "",
            "RESULTADO",
            0,
            "",
            "",
            "RESULTADO",
            0,
            "",
        ],
    ]
    columns = ["C", "D", "E", "F", "G"]
    rows = [2, 3, 4, 5, 8, 9, 10, 13, 14, 15, 18, 19, 22, 23]
    row_titles = [0, 1, 2, 7, 8, 12, 13, 17, 18, 21, 22]
    for column in range(len(columns)):
        for row in range(24):
            sheet[f"{columns[column]}{row+3}"] = cell_values[column][row]
            sheet[f"{columns[column]}{row+3}"].border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            if row in rows and column == 0:
                sheet[f"{columns[column]}{row+3}"].alignment = Alignment(
                    horizontal="left", vertical="center"
                )
            else:
                sheet[f"{columns[column]}{row+3}"].alignment = Alignment(
                    horizontal="center", vertical="center"
                )
            if row in row_titles:
                sheet[f"{columns[column]}{row+3}"].font = Font(bold=True)

    ###     merging cells
    rows = [2, 3, 4, 9, 10, 14, 15, 19, 20, 23, 24, 27]
    for row in rows:
        sheet.merge_cells(start_row=row, end_row=row, start_column=3, end_column=7)

    ###     removing border between topics
    rows = [9, 14, 19, 23]
    for row in rows:
        for column in range(3, 8):
            sheet.cell(row=row, column=column).border = Border()

    ###     getting double borders
    for i in range(3, 8):
        sheet.cell(row=3, column=i).border = Border(
            top=Side(style="double"),
            bottom=Side(style="double"),
            right=Side(style="double"),
            left=Side(style="double"),
        )
    rows = [5, 6, 7, 11, 12, 16, 17, 21, 25]
    for row in range(4, 28):
        for column in range(3, 8):
            if column == 3 and row in rows:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                    left=Side(style="double"),
                    right=Side(style="thin"),
                )
            elif column == 7 and row in rows:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                    left=Side(style="thin"),
                    right=Side(style="double"),
                )
    for row in range(4, 25):
        for column in range(3, 8):
            if row == 4 or row == 10 or row == 15 or row == 20 or row == 24:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="double"),
                    bottom=Side(style="thin"),
                    left=Side(style="double"),
                    right=Side(style="double"),
                )
    for row in range(8, 28):
        for column in range(3, 8):
            if row == 8 or row == 13 or row == 18 or row == 22 or row == 26:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thin"),
                    bottom=Side(style="double"),
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                )
    rows = [8, 13, 18, 22, 26]
    for row in rows:
        sheet.cell(row=row, column=3).border = Border(
            top=Side(style="thin"),
            bottom=Side(style="double"),
            left=Side(style="double"),
            right=Side(style="thin"),
        )
        sheet.cell(row=row, column=7).border = Border(
            top=Side(style="thin"),
            bottom=Side(style="double"),
            left=Side(style="thin"),
            right=Side(style="double"),
        )

    ###     getting thick borders around the page
    for row in range(2, 28):
        for column in range(2, 9):
            if column == 2:
                sheet.cell(row=row, column=column).border = Border(
                    left=Side(style="thick")
                )
            elif column == 8:
                sheet.cell(row=row, column=column).border = Border(
                    right=Side(style="thick")
                )
            elif row == 2:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thick")
                )
            elif row == 27:
                sheet.cell(row=row, column=column).border = Border(
                    bottom=Side(style="thick")
                )
    for row in range(2, 28):
        for column in range(2, 9):
            if row == 2 and column == 2:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thick"), left=Side(style="thick")
                )
            elif row == 2 and column == 8:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thick"), right=Side(style="thick")
                )
            elif row == 27 and column == 2:
                sheet.cell(row=row, column=column).border = Border(
                    bottom=Side(style="thick"), left=Side(style="thick")
                )
            elif row == 27 and column == 8:
                sheet.cell(row=row, column=column).border = Border(
                    bottom=Side(style="thick"), right=Side(style="thick")
                )

    workbook.save(caminho)

def create_sheet_fora_de_prazo_man(caminho):
    locale.setlocale(locale.LC_TIME, "pt_BR")
    today = time.strftime("%B").upper()

    workbook = openpyxl.load_workbook(caminho)
    sheet = workbook.create_sheet("Fora Prazo Man")

    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    ###     setting default row size
    for row in range(1000):
        formating_row(sheet, row, 20)

    ###     setting row sizes
    rows = [2, 3, 4, 9, 14, 19, 23, 27]
    row_sizes = [7, 39.5, 31.5, 7, 7, 7, 7, 7]
    for i in range(len(rows)):
        formating_row(sheet, rows[i], row_sizes[i])

    ###     setting column sizes
    columns = ["B", "C", "D", "E", "F", "G", "H", "I", "J", "K"]
    column_sizes = [1, 8.57, 19.43, 11.14, 27.43, 7, 13.43, 13.43, 13.43, 1]
    for i in range(len(columns)):
        formating_column(sheet, columns[i], column_sizes[i])

    ###     setting cell values
    cell_values = [
        [
            f"SERVIÇOS FORA DE PRAZO MANUTENÇÃO - MÊS DE {today}",
            "EQUIPE"
        ],
        [
            "",
            "LOCAL"
        ],
        [
            "",
            "SERV EXEC"
        ],
        [
            "",
            "DESCRIÇÃO DO SERVIÇO"
        ],
        [
            "",
            "PRAZO"
        ],
        [
            "",
            "DATA/HORA VCTO"
        ],
        [
            "",
            "DATA/HORA FIM EXEC"
        ],
        [
            "",
            "TEMPO FORA DE PRAZO"
        ]   
    ]
    columns = ["C", "D", "E", "F", "G", "H", "I", "J"]
    rows = [2, 3]
    row_titles = [0, 1]
    for column in range(len(columns)):
        for row in range(2):
            sheet[f"{columns[column]}{row+3}"] = cell_values[column][row]
            sheet[f"{columns[column]}{row+3}"].border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            if row in rows and column == 0:
                sheet[f"{columns[column]}{row+3}"].alignment = Alignment(
                    horizontal="left", vertical="center"
                )
            else:
                sheet[f"{columns[column]}{row+3}"].alignment = Alignment(
                    horizontal="center", vertical="center", wrap_text=True
                )
            if row in row_titles:
                sheet[f"{columns[column]}{row+3}"].font = Font(bold=True)

    ###     merging cells
    rows = [2, 3]
    for row in rows:
        sheet.merge_cells(start_row=row, end_row=row, start_column=3, end_column=10)

    ###     getting double borders
    for i in range(3, 11):
        sheet.cell(row=3, column=i).border = Border(
            top=Side(style="double"),
            bottom=Side(style="double"),
            right=Side(style="double"),
            left=Side(style="double"),
        )
    for i in range(3, 11):
        sheet.cell(row=4, column=i).border = Border(
            top=Side(style="double"),
            bottom=Side(style="double"),
            right=Side(style="double"),
            left=Side(style="double"),
        )
    rows = [5, 6]
    for row in range(4, 6):
        for column in range(3, 11):
            if column == 3 and row in rows:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                    left=Side(style="double"),
                    right=Side(style="thin"),
                )
            elif column == 10 and row in rows:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thin"),
                    bottom=Side(style="thin"),
                    left=Side(style="thin"),
                    right=Side(style="double"),
                )
    for row in range(4, 6):
        for column in range(3, 11):
            if row == 4 or row == 10 or row == 15 or row == 20 or row == 24:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="double"),
                    bottom=Side(style="thin"),
                    left=Side(style="double"),
                    right=Side(style="double"),
                )
    for row in range(4, 6):
        for column in range(3, 11):
            if row == 8 or row == 13 or row == 18 or row == 22 or row == 26:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thin"),
                    bottom=Side(style="double"),
                    left=Side(style="thin"),
                    right=Side(style="thin"),
                )
    
    ###     getting thick borders around the page
    for row in range(2, 8):
        for column in range(2, 12):
            if column == 2:
                sheet.cell(row=row, column=column).border = Border(
                    left=Side(style="thick")
                )
            elif column == 12:
                sheet.cell(row=row, column=column).border = Border(
                    right=Side(style="thick")
                )
            elif row == 2:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thick")
                )
            elif row == 7:
                sheet.cell(row=row, column=column).border = Border(
                    bottom=Side(style="thick")
                )
    for row in range(2, 8):
        for column in range(2, 12):
            if row == 2 and column == 2:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thick"), left=Side(style="thick")
                )
            elif row == 2 and column == 12:
                sheet.cell(row=row, column=column).border = Border(
                    top=Side(style="thick"), right=Side(style="thick")
                )
            elif row == 8 and column == 2:
                sheet.cell(row=row, column=column).border = Border(
                    bottom=Side(style="thick"), left=Side(style="thick")
                )
            elif row == 8 and column == 12:
                sheet.cell(row=row, column=column).border = Border(
                    bottom=Side(style="thick"), right=Side(style="thick")
                )

    workbook.save(caminho)
