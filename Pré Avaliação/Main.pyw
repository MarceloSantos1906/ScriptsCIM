import tkinter as tk
from tkinter import filedialog
import pathlib
import FormatarPlanilhas
import threading
import ctypes
import openpyxl
import CriarPlanilha
import ImportarDados


class Window:
    def __init__(self):
        threading.Thread.__init__(self)
        self.path = pathlib.Path(__file__).parent.resolve()
        self.label_qlik = tk.Label(text="Qlik:",
            width=10,
            anchor="e")
        self.label_bdo = tk.Label(text="BDOs:",
            width=10,
            anchor="e")
        self.label_fora_de_prazo = tk.Label(text="Fora de prazo:",
            width=10,
            anchor="e")
        self.label_cef = tk.Label(text="CEF:",
            width=10,
            anchor="e")
        self.label_salvar = tk.Label(text="Salvar em:",
            width=10,
            anchor="e")
        self.label_status = tk.Label(
            text="Ao selecionar as planilhas pela primeira vez clique em formatar",
            font=("Arial", 7)
        )

        self.label_qlik.grid(column=0, row=0, padx=10, pady=(5, 5))
        self.label_bdo.grid(column=0, row=1, padx=10, pady=(0, 5))
        self.label_fora_de_prazo.grid(column=0, row=2, padx=10, pady=(0, 5))
        self.label_cef.grid(column=0, row=3, padx=10, pady=(5, 5))
        self.label_salvar.grid(column=0, row=4, padx=10, pady=(5, 5))
        self.label_status.grid(columnspan=4, row=6, padx=10, pady=(5, 5))

        self.entry_qlik = tk.Entry(width=50)
        self.entry_bdo = tk.Entry(width=50)
        self.entry_fora_de_prazo = tk.Entry(width=50)
        self.entry_cef = tk.Entry(width=50)
        self.entry_salvar = tk.Entry(width=50)

        self.entry_qlik.grid(column=1, row=0, pady=(5, 5))
        self.entry_bdo.grid(column=1, row=1, pady=(0, 5))
        self.entry_fora_de_prazo.grid(column=1, row=2, pady=(0, 5))
        self.entry_cef.grid(column=1, row=3, pady=(0, 5))
        self.entry_salvar.grid(column=1, row=4, pady=(0, 5))

        self.button_escolher_arquivo_qlik = tk.Button(
            text="Escolher", command=lambda: self.open_file(self.entry_qlik)
        )
        self.button_escolher_arquivo_bdo = tk.Button(
            text="Escolher", command=lambda: self.open_file(self.entry_bdo)
        )
        self.button_escolher_arquivo_fora_de_prazo = tk.Button(
            text="Escolher", command=lambda: self.open_file(self.entry_fora_de_prazo)
        )
        self.button_escolher_arquivo_cef = tk.Button(
            text="Escolher", command=lambda: self.open_file(self.entry_cef)
        )

        self.button_formatar_qlik = tk.Button(
            text="Formatar", command=lambda: self.threading(self.FormatarQlik)
        )
        self.button_formatar_bdo = tk.Button(
            text="Formatar", command=lambda: self.threading(self.FormatarBDOs)
        )
        self.button_formatar_fora_de_prazo = tk.Button(
            text="Formatar", command=lambda: self.threading(self.FormatarAtrasados)
        )
        self.button_escolher_salvar = tk.Button(
            text="Escolher", command=lambda: self.save_location(self.entry_salvar)
        )

        self.button_exportar = tk.Button(
            text="Exportar", command=lambda: self.threading(self.Exportar)
        )

        self.button_escolher_arquivo_qlik.grid(column=2, row=0, padx=(10, 0), pady=2)
        self.button_escolher_arquivo_bdo.grid(column=2, row=1, padx=(10, 0), pady=2)
        self.button_escolher_arquivo_fora_de_prazo.grid(column=2, row=2, padx=(10, 0), pady=2)
        self.button_escolher_arquivo_cef.grid(column=2, row=3, padx=(10, 0), pady=2)

        self.button_formatar_qlik.grid(column=3, row=0, padx=(10, 0), pady=2)
        self.button_formatar_bdo.grid(column=3, row=1, padx=(10, 0), pady=2)
        self.button_formatar_fora_de_prazo.grid(column=3, row=2, padx=(10, 0), pady=2)
        self.button_escolher_salvar.grid(column=2, row=4, padx=(10, 0), pady=2)
        self.button_exportar.grid(column=1, columnspan=2, row=5, pady=2, sticky=("EW"))

    def threading(self, target):
        th = threading.Thread(target=target)
        th.start()

    def open_file(self, entry):
        filename = filedialog.askopenfilename(
            initialdir="Downloads",
            title="Select a File",
            filetypes=(("Excel files", "*.xlsx*"), ("all files", "*.*")),
        )
        entry.delete(0, tk.END)
        entry.insert(0, filename)

    def save_location(self, entry):
        filename = filedialog.askdirectory(initialdir="Downloads")
        entry.delete(0, tk.END)
        entry.insert(0, filename)

    def FormatarQlik(self):
        FormatarPlanilhas.FormatarQlik(self.label_status, self.entry_qlik.get())

    def FormatarBDOs(self):
        FormatarPlanilhas.FormatarBDOs(self.label_status, self.entry_bdo.get())

    def FormatarAtrasados(self):
        FormatarPlanilhas.FormatarAtrasados(self.label_status, self.entry_fora_de_prazo.get())

    def Exportar(self):
        workbook = openpyxl.load_workbook(self.entry_qlik.get())
        qlik = workbook.sheetnames[0]
        workbook = openpyxl.load_workbook(self.entry_bdo.get())
        bdos = workbook.sheetnames[0]
        workbook = openpyxl.load_workbook(self.entry_fora_de_prazo.get())
        atrasados = workbook.sheetnames[0]
        if (self.entry_cef.get() != ""):
            workbook = openpyxl.load_workbook(self.entry_cef.get())
            cef = workbook.sheetnames[0]
        Relatorio_workbook = "Relatorio.xlsx"
        CriarPlanilha.create_sheet(f"{self.entry_salvar.get()}/{Relatorio_workbook}")
        workbook = openpyxl.load_workbook(f"{self.entry_salvar.get()}/{Relatorio_workbook}")
        self.label_status.config(text="Copiando Qlik")
        FormatarPlanilhas.copiarPlanilha(
            self.entry_salvar.get(), Relatorio_workbook, "Qlik", self.entry_qlik.get(), qlik
        )
        self.label_status.config(text="Copiando BDOs")
        FormatarPlanilhas.copiarPlanilha(
            self.entry_salvar.get(), Relatorio_workbook, "Relatorio BDO", self.entry_bdo.get(), bdos
        )
        self.label_status.config(text="Copiando Atrasados")
        FormatarPlanilhas.copiarPlanilha(
            self.entry_salvar.get(),
            Relatorio_workbook,
            "Relatorio Atrasados",
            self.entry_fora_de_prazo.get(),
            atrasados,
        )
        print(f"CEF: '{len(self.entry_cef.get())}'")
        if (len(self.entry_cef.get()) > 1):
            self.label_status.config(text="Copiando CEF")
            FormatarPlanilhas.copiarPlanilha(
                self.entry_salvar.get(), Relatorio_workbook, "Relatorio CEF", self.entry_cef.get(), cef
            )
        ImportarDados.import_data(
            caminho=self.entry_salvar.get(),
            workbook_Relatorio=Relatorio_workbook,
            label=self.label_status
        )
        self.label_status.config(text=f"Salvo em {self.entry_salvar.get()}/{Relatorio_workbook}")

    root = tk.Tk()
    root.minsize(530, 210)
    root.maxsize(530, 210)


class thread_with_exception(threading.Thread):
    def __init__(self, name):
        threading.Thread.__init__(self)
        self.name = name

    def get_id(self):
        if hasattr(self, "_thread_id"):
            return self._thread_id
        for id, thread in threading._active.items():
            if thread is self:
                return id

    def raise_exception(self):
        thread_id = self.get_id()
        res = ctypes.pythonapi.PyThreadState_SetAsyncExc(
            thread_id, ctypes.py_object(SystemExit)
        )
        if res > 1:
            ctypes.pythonapi.PyThreadState_SetAsyncExc(thread_id, 0)

try:
    window = Window()
    window.root.mainloop()
except Exception as e:
    print(f"Exception: {e}")
