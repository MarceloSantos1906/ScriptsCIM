import openpyxl
from datetime import *


def percentage(part, whole):
    if whole == 0:
        return "Div/0"
    return float(part) / float(whole)


def atende_atparcial_natende(sheet, row, min_percent, max_percent):
    if sheet[f"F{row}"].value == "Div/0":
        return "ERROR"
    elif sheet[f"F{row}"].value is not None:
        if float(sheet[f"F{row}"].value) <= min_percent:
            return "Atende"
        elif float(sheet[f"F{row}"].value) <= max_percent:
            return "Atende Parcial"
        elif float(sheet[f"F{row}"].value) <= max_percent:
            return "Não Atende"
        else:
            return "ERROR"
    else:
        return "ERROR"


def fomat_data(sheet):
    ###     setting percentages
    rows = [6, 7, 8, 12, 13, 17, 18, 22, 26]
    for row in rows:
        sheet[f"F{row}"] = percentage(sheet[f"D{row}"].value, sheet[f"E{row}"].value)
        if not (sheet[f"F{row}"] == "Div/0"):
            sheet[f"F{row}"].number_format = "0.00%"

    ###     settings atende / atende parcial / n atende
    rows = [6, 7, 8, 12, 13, 17, 18, 22, 26]
    min_percentages = [1, 1, 2.5, 3, 1]
    max_percentages = [2, 1.5, 3.5, 6, 1.5]
    for row in range(len(rows)):
        if sheet[f"F{rows[row]}"].value == "Div/0":
            sheet[f"G{rows[row]}"] = "ERROR"
            continue
        sheet[f"G{rows[row]}"] = atende_atparcial_natende(
            sheet=sheet,
            row=rows[row],
            min_percent=min_percentages[row],
            max_percent=max_percentages[row],
        )


def count_qlik(label, sheetQlik):
    obras = 0
    servicos_man = 0
    servicos_pav = 0
    servicos_22_07 = 0
    protocolos_22_07hrs = []
    for row in sheetQlik.iter_rows(
        min_row=2, max_row=sheetQlik.max_row, min_col=1, max_col=1
    ):
        for cell in row:
            label.config(
                text=f"Contando Qlik, linha: {cell.row} de {sheetQlik.max_row}"
            )
            if sheetQlik[f"C{cell.row}"].value is not None:
                if ((int(sheetQlik[f"C{cell.row}"].value) >= 2200) and (int(sheetQlik[f"C{cell.row}"].value) <= 2359)) or (
                    (int(sheetQlik[f"C{cell.row}"].value) >= 0) and (int(sheetQlik[f"C{cell.row}"].value) <= 700) and ((str(sheetQlik[f"A{cell.row}"].value) not in protocolos_22_07hrs))
                ):
                    servicos_22_07 += 1
                    protocolos_22_07hrs.append(str(sheetQlik[f"A{cell.row}"].value))
                elif not ((str(sheetQlik[f"O{cell.row}"].value) in servicoPavimento) or (str(sheetQlik[f"O{cell.row}"].value) in codigo_obras)):
                    servicos_man += 1
                elif str(sheetQlik[f"O{cell.row}"].value) in servicoPavimento:
                    servicos_pav += 1
                elif str(sheetQlik[f"O{cell.row}"].value) in codigo_obras:
                    obras += 1
    return servicos_man, servicos_pav, servicos_22_07, protocolos_22_07hrs, obras

def count_cef(label, sheetCef):
    cef_manutencao = 0
    cef_pavimento = 0
    for row in sheetCef.iter_rows(
        min_row=2, max_row=sheetCef.max_row, min_col=1, max_col=1
    ):
        for cell in row:
            label.config(
                text=f"Contando Cef, linha: {cell.row} de {sheetCef.max_row}"
            )
            if not ((str(sheetCef[f"H{cell.row}"].value) in servicoPavimento)):
                cef_manutencao += 1
            elif str(sheetCef[f"H{cell.row}"].value) in servicoPavimento:
                cef_pavimento += 1
    return cef_manutencao, cef_pavimento


def count_servicos_bdo(label, sheetBDO, subItem: str):
    servicos = 0
    equipe = []
    enderecos = []
    observacao = []
    cod_serviço = []
    for row in sheetBDO.iter_rows(
        min_row=4, max_row=sheetBDO.max_row, min_col=1, max_col=1
    ):
        for cell in row:
            label.config(text=f"Contando BDOs, linha: {cell.row} de {sheetBDO.max_row}")
            if (
                (str(sheetBDO[f"C{cell.row}"].value).lower() == "cef")
                and (str(sheetBDO[f"S{cell.row}"].value).lower() == "não")
                and subItem in (str(sheetBDO[f"U{cell.row}"].value).lower())
            ):
                servicos += 1
                equipe.append(str(sheetBDO[f"M{cell.row}"].value))
                enderecos.append(str(sheetBDO[f"N{cell.row}"].value))
                observacao.append(str(sheetBDO[f"O{cell.row}"].value))
                cod_serviço.append(int(sheetBDO[f"K{cell.row}"].value))
    return servicos, equipe, enderecos, observacao, cod_serviço


def count_servicos_atraso(label, sheetAtrasados, protocolos_22_07hrs):
    fora_de_prazo_obras = 0
    fora_prazo_servicos_man = 0
    fora_prazo_servicos_pav = 0
    fora_prazo_servicos_22_07 = 0
    fora_prazo_equipes22_7, fora_prazo_cod_servico22_7, fora_prazo_horario_protoc22_7, fora_prazo_horario_fim_exec22_7, fora_prazo_prazo_exec22_7, fora_prazo_equipes_man, fora_prazo_cod_servico_man, fora_prazo_horario_protoc_man, fora_prazo_horario_fim_exec_man, fora_prazo_prazo_exec_man, fora_prazo_equipes_pav, fora_prazo_cod_servico_pav, fora_prazo_horario_protoc_pav, fora_prazo_horario_fim_exec_pav, fora_prazo_prazo_exec_pav, fora_prazo_equipes_obra, fora_prazo_cod_servico_obra, fora_prazo_horario_protoc_obra, fora_prazo_horario_fim_exec_obra, fora_prazo_prazo_exec_obra = [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], [], []
    for row in sheetAtrasados.iter_rows(
        min_row=4, max_row=sheetAtrasados.max_row, min_col=1, max_col=1
    ):
        for cell in row:
            label.config(
                text=f"Contando Atrasados, linha: {cell.row} de {sheetAtrasados.max_row}"
            )
            if (
                sheetAtrasados[f"G{cell.row}"].value is not None
                and (str(sheetAtrasados[f"V{cell.row}"].value) == "Não")
                and str(sheetAtrasados[f"D{cell.row}"].value) in protocolos_22_07hrs
            ):
                fora_prazo_servicos_22_07 += 1
                fora_prazo_equipes22_7.append(sheetAtrasados[f"R{cell.row}"].value)
                fora_prazo_cod_servico22_7.append(sheetAtrasados[f"G{cell.row}"].value)
                fora_prazo_horario_protoc22_7.append(sheetAtrasados[f"I{cell.row}"].value)
                fora_prazo_horario_fim_exec22_7.append(sheetAtrasados[f"L{cell.row}"].value)
                fora_prazo_prazo_exec22_7.append(sheetAtrasados[f"M{cell.row}"].value)
            elif (
                sheetAtrasados[f"G{cell.row}"].value is not None
                and not (str(sheetAtrasados[f"G{cell.row}"].value) in servicoPavimento)
                and not (str(sheetAtrasados[f"G{cell.row}"].value) in codigo_obras)
                and (str(sheetAtrasados[f"V{cell.row}"].value) == "Não")
            ):
                fora_prazo_servicos_man += 1
                fora_prazo_equipes_man.append(sheetAtrasados[f"R{cell.row}"].value)
                fora_prazo_cod_servico_man.append(sheetAtrasados[f"G{cell.row}"].value)
                fora_prazo_horario_protoc_man.append(sheetAtrasados[f"I{cell.row}"].value)
                fora_prazo_horario_fim_exec_man.append(sheetAtrasados[f"L{cell.row}"].value)
                fora_prazo_prazo_exec_man.append(sheetAtrasados[f"M{cell.row}"].value)
            elif (
                sheetAtrasados[f"G{cell.row}"].value is not None
                and (str(sheetAtrasados[f"G{cell.row}"].value) in servicoPavimento)
                and (str(sheetAtrasados[f"V{cell.row}"].value) == "Não")
            ):
                fora_prazo_servicos_pav += 1
                fora_prazo_equipes_pav.append(sheetAtrasados[f"R{cell.row}"].value)
                fora_prazo_cod_servico_pav.append(sheetAtrasados[f"G{cell.row}"].value)
                fora_prazo_horario_protoc_pav.append(sheetAtrasados[f"I{cell.row}"].value)
                fora_prazo_horario_fim_exec_pav.append(sheetAtrasados[f"L{cell.row}"].value)
                fora_prazo_prazo_exec_pav.append(sheetAtrasados[f"M{cell.row}"].value)
            elif (
                sheetAtrasados[f"G{cell.row}"].value is not None
                and (str(sheetAtrasados[f"G{cell.row}"].value) in codigo_obras)
                and (str(sheetAtrasados[f"V{cell.row}"].value) == "Não")
            ):
                fora_de_prazo_obras += 1
                fora_prazo_equipes_obra.append(sheetAtrasados[f"R{cell.row}"].value)
                fora_prazo_cod_servico_obra.append(sheetAtrasados[f"G{cell.row}"].value)
                fora_prazo_horario_protoc_obra.append(sheetAtrasados[f"I{cell.row}"].value)
                fora_prazo_horario_fim_exec_obra.append(sheetAtrasados[f"L{cell.row}"].value)
                fora_prazo_prazo_exec_obra.append(sheetAtrasados[f"M{cell.row}"].value)
    return fora_prazo_servicos_man, fora_prazo_servicos_pav, fora_prazo_servicos_22_07, fora_de_prazo_obras, fora_prazo_equipes22_7, fora_prazo_cod_servico22_7, fora_prazo_horario_protoc22_7, fora_prazo_horario_fim_exec22_7, fora_prazo_prazo_exec22_7, fora_prazo_equipes_man, fora_prazo_cod_servico_man, fora_prazo_horario_protoc_man, fora_prazo_horario_fim_exec_man, fora_prazo_prazo_exec_man, fora_prazo_equipes_pav, fora_prazo_cod_servico_pav, fora_prazo_horario_protoc_pav, fora_prazo_horario_fim_exec_pav, fora_prazo_prazo_exec_pav, fora_prazo_equipes_obra, fora_prazo_cod_servico_obra, fora_prazo_horario_protoc_obra, fora_prazo_horario_fim_exec_obra, fora_prazo_prazo_exec_obra


def import_data(caminho, workbook_Relatorio, label):
    workbook = openpyxl.load_workbook(f"{caminho}/{workbook_Relatorio}")
    sheets = workbook.sheetnames
    sheetRelatorio = workbook[sheets[0]]
    sheetQlik = workbook[sheets[1]]
    sheetBDO = workbook[sheets[2]]
    sheetAtraso = workbook[sheets[3]]
    print(f"Sheets: {sheets}")
    if "Relatorio CEF" in sheets:
        sheetCef = workbook[sheets[4]]
    (
        servicos_man,
        servicos_pav,
        servicos_22_07,
        protocolos_22_07hrs, obras
    ) = count_qlik(label=label, sheetQlik=sheetQlik)
    if "Relatorio CEF" in sheets:
        (
            cef_manutencao,
            cef_pavimento
        ) = count_cef(label=label, sheetCef=sheetCef)
    retrabalho_manutencao, equipes_ret_man, end_ret_man, obs_ret_man, cod_serv_ret_man= count_servicos_bdo(
        label=label,
        sheetBDO=sheetBDO,
        subItem=("retrabalho sgm"),
    )
    retrabalho_pavimento, equipes_ret_pav, end_ret_pav, obs_ret_pav, cod_serv_ret_pav = count_servicos_bdo(
        label=label,
        sheetBDO=sheetBDO,
        subItem=("retrabalho pavimentos"),
    )
    conformidade_manutencao, equipes_conf_man, end_conf_man, obs_conf_man, cod_serv_conf_man = count_servicos_bdo(
        label=label,
        sheetBDO=sheetBDO,
        subItem=("conformidade sgm"),
    )
    conformidade_pavimento, equipes_conf_pav, end_conf_pav, obs_conf_pav, cod_serv_conf_pav = count_servicos_bdo(
        label=label,
        sheetBDO=sheetBDO,
        subItem=("conformidade pavimentos"),
    )
    fora_prazo_servicos_man, fora_prazo_servicos_pav, fora_prazo_servicos_22_07, fora_de_prazo_obras, fora_prazo_equipes22_7, fora_prazo_cod_servico22_7, fora_prazo_horario_protoc22_7, fora_prazo_horario_fim_exec22_7, fora_prazo_prazo_exec22_7, fora_prazo_equipes_man, fora_prazo_cod_servico_man, fora_prazo_horario_protoc_man, fora_prazo_horario_fim_exec_man, fora_prazo_prazo_exec_man, fora_prazo_equipes_pav, fora_prazo_cod_servico_pav, fora_prazo_horario_protoc_pav, fora_prazo_horario_fim_exec_pav, fora_prazo_prazo_exec_pav, fora_prazo_equipes_obra, fora_prazo_cod_servico_obra, fora_prazo_horario_protoc_obra, fora_prazo_horario_fim_exec_obra, fora_prazo_prazo_exec_obra = count_servicos_atraso(
        label=label, sheetAtrasados=sheetAtraso, protocolos_22_07hrs=protocolos_22_07hrs
    )
    
    sheetRelatorio["D6"] = servicos_man
    sheetRelatorio["D7"] = servicos_pav
    sheetRelatorio["D8"] = servicos_22_07

    sheetRelatorio["E6"] = fora_prazo_servicos_man
    sheetRelatorio["E7"] = fora_prazo_servicos_pav
    sheetRelatorio["D8"] = fora_prazo_servicos_22_07

    sheetRelatorio["F6"] = percentage(sheetRelatorio["E6"].value, sheetRelatorio["D6"].value)
    sheetRelatorio["F6"].number_format = '0.00%'
    sheetRelatorio["F7"] = percentage(sheetRelatorio["E7"].value, sheetRelatorio["D7"].value)
    sheetRelatorio["F7"].number_format = '0.00%'
    sheetRelatorio["F8"] = percentage(sheetRelatorio["E8"].value, sheetRelatorio["D8"].value)
    sheetRelatorio["F8"].number_format = '0.00%'

    sheetRelatorio["G6"] = atende_atparcial_natende(sheet = sheetRelatorio, row = 6, min_percent = 1, max_percent = 2)
    sheetRelatorio["G7"] = atende_atparcial_natende(sheet = sheetRelatorio, row = 7, min_percent = 1, max_percent = 2)
    sheetRelatorio["G8"] = atende_atparcial_natende(sheet = sheetRelatorio, row = 8, min_percent = 1, max_percent = 2)

    if "Relatorio CEF" in sheets:
        sheetRelatorio["D12"] = cef_manutencao
        sheetRelatorio["D13"] = cef_pavimento
    else:
        sheetRelatorio["D12"] = 0
        sheetRelatorio["D13"] = 0

    sheetRelatorio["E12"] = retrabalho_manutencao
    sheetRelatorio["E13"] = retrabalho_pavimento

    sheetRelatorio["F12"] = percentage(sheetRelatorio["E12"].value, sheetRelatorio["D12"].value)
    sheetRelatorio["F12"].number_format = '0.00%'
    sheetRelatorio["F13"] = percentage(sheetRelatorio["E13"].value, sheetRelatorio["D13"].value)
    sheetRelatorio["F13"].number_format = '0.00%'

    sheetRelatorio["G12"] = atende_atparcial_natende(sheet = sheetRelatorio, row = 12, min_percent = 1, max_percent = 1.5)
    sheetRelatorio["G13"] = atende_atparcial_natende(sheet = sheetRelatorio, row = 13, min_percent = 1, max_percent = 1.5)

    if "Relatorio CEF" in sheets:
        sheetRelatorio["D17"] = cef_manutencao
        sheetRelatorio["D18"] = cef_pavimento
    else:
        sheetRelatorio["D17"] = 0
        sheetRelatorio["D18"] = 0

    sheetRelatorio["E17"] = conformidade_manutencao
    sheetRelatorio["E18"] = conformidade_pavimento

    sheetRelatorio["F17"] = percentage(sheetRelatorio["E17"].value, sheetRelatorio["D17"].value)
    sheetRelatorio["F17"].number_format = '0.00%'
    sheetRelatorio["F18"] = percentage(sheetRelatorio["E18"].value, sheetRelatorio["D18"].value)
    sheetRelatorio["F18"].number_format = '0.00%'

    sheetRelatorio["G17"] = atende_atparcial_natende(sheet = sheetRelatorio, row = 17, min_percent = 2.5, max_percent = 3.5)
    sheetRelatorio["G18"] = atende_atparcial_natende(sheet = sheetRelatorio, row = 18, min_percent = 2.5, max_percent = 3.5)

    sheetRelatorio["D22"] = obras
    sheetRelatorio["E22"] = fora_de_prazo_obras

    sheetRelatorio["F22"] = percentage(sheetRelatorio["E22"].value, sheetRelatorio["D22"].value)
    sheetRelatorio["F22"].number_format = '0.00%'

    sheetRelatorio["G22"] = atende_atparcial_natende(sheet = sheetRelatorio, row = 22, min_percent = 3, max_percent = 6)

    sheetRelatorio["D26"] = obras
    sheetRelatorio["E26"] = 0

    sheetRelatorio["F26"] = percentage(sheetRelatorio["E26"].value, sheetRelatorio["D26"].value)
    sheetRelatorio["F26"].number_format = '0.00%'

    sheetRelatorio["G26"] = atende_atparcial_natende(sheet = sheetRelatorio, row = 26, min_percent = 1, max_percent = 1.5)

    workbook.save(f"{caminho}/{workbook_Relatorio}")

def import_fora_prazo_man(caminho, workbook_Relatorio, label):
    workbook = openpyxl.load_workbook(f"{caminho}/{workbook_Relatorio}")
    sheets = workbook.sheetnames
    sheetQlik = workbook[sheets[1]]
    sheetAtraso = workbook[sheets[3]]
    sheet_Fora_Prazo_Manu = workbook[sheets[5]]
    (
        servicos_man,
        servicos_pav,
        servicos_22_07,
        protocolos_22_07hrs, obras
    ) = count_qlik(label=label, sheetQlik=sheetQlik)

    servicos_man, servicos_pav, servicos_22_07, fora_de_prazo_obras, equipes22_7, cod_servico22_7, horario_protoc22_7, horario_fim_exec22_7, prazo_exec22_7, equipes_man, cod_servico_man, horario_protoc_man, horario_fim_exec_man, prazo_exec_man, equipes_pav, cod_servico_pav, horario_protoc_pav, horario_fim_exec_pav, prazo_exec_pav, fora_prazo_cod_servico_obra, cod_servico_obra, horario_protoc_obra, horario_fim_exec_obra, prazo_exec_obra = count_servicos_atraso(
        label=label, sheetAtrasados=sheetAtraso, protocolos_22_07hrs=protocolos_22_07hrs
    )
    
    for row in range(len(equipes_man)):
        sheet_Fora_Prazo_Manu[f"C{row+4}"] = equipes_man[row]
        sheet_Fora_Prazo_Manu[f"E{row+4}"] = cod_servico_man[row]
        sheet_Fora_Prazo_Manu[f"F{row+4}"] = prazo_exec_man[row]
        sheet_Fora_Prazo_Manu[f"G{row+4}"].number_format = "dd-mm-yyyy hh:mm:ss"
        sheet_Fora_Prazo_Manu[f"G{row+4}"] = horario_protoc_man[row]
        sheet_Fora_Prazo_Manu[f"H{row+4}"].number_format = "dd-mm-yyyy hh:mm:ss"
        sheet_Fora_Prazo_Manu[f"H{row+4}"] = horario_fim_exec_man[row]
        sheet_Fora_Prazo_Manu[f"I{row+4}"].number_format = "dd-mm-yyyy hh:mm:ss"
        sheet_Fora_Prazo_Manu[f"I{row+4}"] = f"=H{row} - G{row}"

servicoPavimento = ["1555", "1556", "1557", "1560"]
codigo_obras = ["0746", "0755", "1670", "1920", "3065", "3500", "3520"]
baseUniao = [
    "088",
    "104",
    "196",
    "211",
    "283",
    "400",
    "644",
    "645",
    "668",
    "674",
    "682",
]
baseBituruna = ["055", "413"]
